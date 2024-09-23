from openpyxl.chart import Reference, BarChart
from openpyxl.reader.excel import load_workbook

def correct_prices_in_the_sheet(filename: str, updated_filename: str) :
    wb = load_workbook(filename)
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row+1):
        cell = sheet.cell(row,3)
        print(cell.value)
        correct_price = float(cell.value) * 0.9
        correct_price_cell = sheet.cell(row,4)
        correct_price_cell.value = round(correct_price,2)
        print(correct_price_cell.value)
    corrected_values = Reference(sheet,
              min_row=2,
              max_row=4,
              min_col=4,
              max_col=4)
    chart = BarChart()
    chart.add_data(corrected_values,'e2')
    sheet.add_chart(chart)

    wb.save(updated_filename)

#correct_prices_in_the_sheet('../datadocs/price_list.xlsx','datadocs/price_list_updated_v1.xlsx')

#sheet = wb['MySheet']
#print(sheet)
#cell = sheet['b2']
#print(cell.value)
#cell = sheet.cell(4,1)
#print(cell.value)

